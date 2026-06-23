"""XLSX extractor: per-sheet row blocks with repeated header; sheet + row-range location"""

import hashlib
import io
from collections.abc import Iterable

import openpyxl

from app.core.errors import ExtractionError
from app.ingestion.detector import register_extractor
from app.ingestion.models import Element, ElementType, ExtractedDocument, Location

_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
_ROWS_PER_BLOCK = 50


def _row_text(values: Iterable[object]) -> str:
    return " | ".join("" if v is None else str(v) for v in values)


def extract_xlsx(filename: str, data: bytes) -> ExtractedDocument:
    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as exc:
        raise ExtractionError(f"Failed to open XLSX: {exc}") from exc

    elements: list[Element] = []
    try:
        num_sheets = len(wb.sheetnames)
        for ws in wb.worksheets:
            rows = [
                (idx, vals)
                for idx, vals in enumerate(ws.iter_rows(values_only=True), start=1)
                if any(v is not None for v in vals)
            ]
            if not rows:
                continue
            header = _row_text(rows[0][1])
            data_rows = rows[1:]
            for i in range(0, len(data_rows), _ROWS_PER_BLOCK):
                block = data_rows[i : i + _ROWS_PER_BLOCK]
                body = "\n".join(_row_text(vals) for _, vals in block)
                elements.append(
                    Element(
                        text=f"{header}\n{body}",
                        element_type=ElementType.CELL_BLOCK,
                        location=Location(
                            sheet=ws.title, row_start=block[0][0], row_end=block[-1][0]
                        ),
                    )
                )
    finally:
        wb.close()

    if not elements:
        raise ExtractionError("XLSX has no extractable data")

    return ExtractedDocument(
        filename=filename,
        mime_type=_MIME,
        format="xlsx",
        file_hash=hashlib.sha256(data).hexdigest(),
        size_bytes=len(data),
        elements=elements,
        num_units=num_sheets,
    )


register_extractor("xlsx", extract_xlsx)
